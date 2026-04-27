#!/usr/bin/env python3
"""Build Instagram catalog (kategori/varian/produk) into an Excel table with embedded images.

Usage example:
python tools/ig_catalog/build_catalog.py --accounts kasyaraa.co kasyaraa.catalog --max-posts 150
"""

from __future__ import annotations

import argparse
import os
import re
import textwrap
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import urlparse

import instaloader
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font


CATEGORY_RULES: dict[str, list[str]] = {
    "dress": ["dress", "gamis", "abaya", "tunik", "tunik"],
    "set": ["set", "setelan", "one set", "2in1", "2 in 1", "twinset"],
    "atasan": ["blouse", "kemeja", "shirt", "top", "atasan", "outer", "cardigan"],
    "bawahan": ["celana", "rok", "pants", "skirt", "kulot", "legging"],
    "hijab": ["hijab", "pashmina", "bergo", "khimar", "scarf"],
    "aksesoris": ["tas", "bag", "sepatu", "sandal", "belt", "bros", "pin", "kaos kaki"],
}

VARIANT_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"\b(xs|s|m|l|xl|xxl|xxxl)\b", re.IGNORECASE),
    re.compile(r"\bsize\s*[:\-]?\s*([a-z0-9, /-]+)", re.IGNORECASE),
    re.compile(r"\bwarna\s*[:\-]?\s*([a-z0-9, /-]+)", re.IGNORECASE),
    re.compile(r"\bcolor\s*[:\-]?\s*([a-z0-9, /-]+)", re.IGNORECASE),
    re.compile(r"\bvarian\s*[:\-]?\s*([a-z0-9, /-]+)", re.IGNORECASE),
]


@dataclass
class CatalogRow:
    akun: str
    shortcode: str
    tanggal: str
    url_post: str
    produk: str
    kategori: str
    varian: str
    caption: str
    image_path: Path | None


def normalize_account(account_or_url: str) -> str:
    value = account_or_url.strip()
    if value.startswith("http://") or value.startswith("https://"):
        parsed = urlparse(value)
        parts = [part for part in parsed.path.split("/") if part]
        if parts:
            value = parts[0]
    return value.lstrip("@")


def clean_caption(text: str) -> str:
    text = text.replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def infer_category(caption: str) -> str:
    low = caption.lower()
    scores: dict[str, int] = {}
    for category, keywords in CATEGORY_RULES.items():
        score = sum(1 for keyword in keywords if keyword in low)
        if score > 0:
            scores[category] = score

    if not scores:
        return "lainnya"
    return sorted(scores.items(), key=lambda x: x[1], reverse=True)[0][0]


def infer_variant(caption: str) -> str:
    hits: list[str] = []
    for pattern in VARIANT_PATTERNS:
        for match in pattern.finditer(caption):
            if match.lastindex:
                hits.append(match.group(1).strip())
            else:
                hits.append(match.group(0).strip())

    if not hits:
        return "-"

    normalized: list[str] = []
    seen: set[str] = set()
    for item in hits:
        item_clean = re.sub(r"\s+", " ", item).strip(" ,.-")
        key = item_clean.lower()
        if key and key not in seen:
            normalized.append(item_clean)
            seen.add(key)

    return " | ".join(normalized[:5]) if normalized else "-"


def infer_product_name(caption: str, fallback: str) -> str:
    lines = [line.strip() for line in caption.splitlines() if line.strip()]
    for line in lines:
        if line.startswith("#"):
            continue
        # drop usernames / promo lines early
        if line.startswith("@"):
            continue
        cleaned = re.sub(r"https?://\S+", "", line).strip()
        if len(cleaned) >= 3:
            return textwrap.shorten(cleaned, width=80, placeholder="...")
    return fallback


def maybe_login(
    loader: instaloader.Instaloader,
    cli_user: str | None = None,
    cli_pass: str | None = None,
    cli_sessionid: str | None = None,
) -> None:
    if cli_sessionid:
        loader.context._session.cookies.set(  # noqa: SLF001
            "sessionid", cli_sessionid, domain=".instagram.com"
        )
        return

    username = cli_user or os.getenv("IG_USER")
    password = cli_pass or os.getenv("IG_PASS")
    if username and password:
        loader.login(username, password)


def download_image(url: str, dest: Path) -> Path | None:
    try:
        response = requests.get(url, timeout=45)
        response.raise_for_status()
    except requests.RequestException:
        return None

    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(response.content)
    return dest


def fetch_profile_posts(
    loader: instaloader.Instaloader,
    account: str,
    max_posts: int,
    image_dir: Path,
) -> Iterable[CatalogRow]:
    profile = instaloader.Profile.from_username(loader.context, account)

    count = 0
    for post in profile.get_posts():
        if count >= max_posts:
            break

        caption = clean_caption(post.caption or "")
        shortcode = post.shortcode
        post_url = f"https://www.instagram.com/p/{shortcode}/"
        image_file = image_dir / account / f"{shortcode}.jpg"
        image_path = download_image(post.url, image_file)

        yield CatalogRow(
            akun=account,
            shortcode=shortcode,
            tanggal=post.date_utc.strftime("%Y-%m-%d %H:%M:%S"),
            url_post=post_url,
            produk=infer_product_name(caption, fallback=shortcode),
            kategori=infer_category(caption),
            varian=infer_variant(caption),
            caption=caption,
            image_path=image_path,
        )

        count += 1


def write_excel(rows: list[CatalogRow], output_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "katalog"

    headers = [
        "no",
        "gambar",
        "akun",
        "shortcode",
        "tanggal",
        "produk",
        "kategori",
        "varian",
        "url_post",
        "caption",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font

    for idx, row in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=3, value=row.akun)
        ws.cell(row=idx, column=4, value=row.shortcode)
        ws.cell(row=idx, column=5, value=row.tanggal)
        ws.cell(row=idx, column=6, value=row.produk)
        ws.cell(row=idx, column=7, value=row.kategori)
        ws.cell(row=idx, column=8, value=row.varian)
        ws.cell(row=idx, column=9, value=row.url_post)
        ws.cell(row=idx, column=10, value=row.caption)

        if row.image_path and row.image_path.exists():
            try:
                img = XLImage(str(row.image_path))
                img.width = 96
                img.height = 96
                ws.add_image(img, f"B{idx}")
            except Exception:
                pass

        ws.row_dimensions[idx].height = 78

    # Column widths
    widths = {
        "A": 6,
        "B": 16,
        "C": 20,
        "D": 16,
        "E": 20,
        "F": 42,
        "G": 16,
        "H": 28,
        "I": 44,
        "J": 90,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build Instagram catalog Excel with product classification and embedded images.")
    parser.add_argument(
        "--accounts",
        nargs="+",
        required=True,
        help="Instagram usernames, ex: kasyaraa.co kasyaraa.catalog",
    )
    parser.add_argument("--max-posts", type=int, default=120, help="Max posts per account")
    parser.add_argument(
        "--output",
        default="tools/ig_catalog/output/katalog_instagram.xlsx",
        help="Output Excel path",
    )
    parser.add_argument(
        "--image-dir",
        default="tools/ig_catalog/output/images",
        help="Local folder for downloaded images",
    )
    parser.add_argument(
        "--sessionid",
        default=None,
        help="Instagram sessionid cookie (optional, use this to bypass repeated checkpoint).",
    )
    parser.add_argument("--ig-user", default=None, help="Instagram username login (optional)")
    parser.add_argument("--ig-pass", default=None, help="Instagram password login (optional)")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = Path(args.output).resolve()
    image_dir = Path(args.image_dir).resolve()

    loader = instaloader.Instaloader(
        download_pictures=False,
        download_videos=False,
        download_video_thumbnails=False,
        download_comments=False,
        save_metadata=False,
        compress_json=False,
        quiet=True,
    )

    maybe_login(
        loader,
        cli_user=args.ig_user,
        cli_pass=args.ig_pass,
        cli_sessionid=args.sessionid,
    )

    rows: list[CatalogRow] = []
    for raw_account in args.accounts:
        account = normalize_account(raw_account)
        try:
            rows.extend(
                list(fetch_profile_posts(loader, account=account, max_posts=args.max_posts, image_dir=image_dir))
            )
            print(f"[OK] {account}")
        except instaloader.exceptions.ProfileNotExistsException:
            print(f"[SKIP] {account} tidak ditemukan.")
        except instaloader.exceptions.QueryReturnedForbiddenException:
            print(
                f"[SKIP] {account} kena 403 dari Instagram. "
                "Coba login dulu: set IG_USER & IG_PASS."
            )
        except instaloader.exceptions.ConnectionException as exc:
            print(f"[SKIP] {account} gagal koneksi: {exc}")

    # Sort newest first
    rows.sort(key=lambda item: datetime.fromisoformat(item.tanggal), reverse=True)
    write_excel(rows, output_path)

    print(f"OK. Total baris: {len(rows)}")
    print(f"Excel: {output_path}")


if __name__ == "__main__":
    main()
